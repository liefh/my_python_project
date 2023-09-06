import datetime
import logging
import time

import openpyxl
import pymysql
from openpyxl.styles import Font, Alignment, PatternFill,colors,Border,Side
import oss2
import requests

import os
import sys

from typing import List
#
# from alibabacloud_dingtalk.robot_1_0.client import Client as dingtalkrobot_1_0Client
# from alibabacloud_tea_openapi import models as open_api_models
# from alibabacloud_dingtalk.robot_1_0 import models as dingtalkrobot__1__0_models
# from alibabacloud_tea_util import models as util_models
# from alibabacloud_tea_util.client import Client as UtilClient

logging.basicConfig(
    # filename='xxx',
    # filemode='a',
    # encoding='utf-8',
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s %(message)s %(filename)s %(funcName)s[line:%(lineno)d]',
    datefmt="%FT%T",
)
webhook_url='https://oapi.dingtalk.com/robot/send?access_token=58af4c18b06991e6928b6d948ff7583e9383f730d777c24d64ba94d58b3dbcab'
database_user = 'ipfs'
database_password = 'Ipfs@123'
database_host = 'rm-bp1vf78w1utqcer11do.mysql.rds.aliyuncs.com'
database_name = 'monitor'


work_sheet = '运维交接表'
row_index = 1
col_index = 1
col_offset = 3

bad_increated_sql = 'SELECT display_for_metabase.name as 集群名称,miner as 集群ID,round(if(today_increase=0,0,today_increase/expect-truncate(TIMESTAMPDIFF(MINUTE,CURRENT_DATE,NOW())/1440,2))*100,2) AS "今日差距" from display_for_metabase left join  cluster_list on display_for_metabase.miner=cluster_list.f0 where if(today_increase=0,0,today_increase/expect-truncate(TIMESTAMPDIFF(MINUTE,CURRENT_DATE,NOW())/1440,2))<-0.06;'

fault_sectors_sql = "select name as '集群名称',f0 as '集群ID',faults*size/1024 as '惩罚算力' from cluster_list where faults>0 order by faults desc;"

orphan_block_sql = "select cluster AS 集群名称,miner AS '集群ID',local_block_list AS 孤块高度,reason AS 孤块原因 from filecoin_cluster_orphan_block where datetime>DATE_FORMAT(NOW(),'%Y-%m-%d 00:00:00') and venus_chain NOT IN ('incubator');"

access_key_id = 'LTAI5tQxr1wiiX7Rm9pSANvq'
access_key_secret = 'BCOwbbdZgbsTIOmjKPGIHdOSEyZjaK'
# endpoint = 'http://oss-accelerate.aliyuncs.com'
endpoint_name = 'oss-accelerate.aliyuncs.com'
bucket_name = 'ipfs-actor-cluster'

def mysql_query_data(sql):
    try:
        conn = pymysql.connect(host=database_host, user=database_user, password=database_password, db=database_name)
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(sql)
        data = cursor.fetchall()
        logging.debug('成功连接数据库')
        return data
    except Exception as e:
        logging.error(f'连接mysql报错 : {e}')
    finally:
        cursor.close()
        conn.close()

def get_duty_name():
    duty_date = time.strftime('%Y年%#m月%#d日', time.localtime(time.time()))
    if mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';") is not None:
        duty_name=mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';")
        print('1'*10,duty_name)
        return duty_name[0]['duty_name']

    duty_date = time.strftime('%Y年%m月%d日', time.localtime(time.time()))
    if mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';") is not None:
        duty_name=mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';")
        return duty_name[0]['duty_name']


    duty_date = time.strftime('%Y年%#m月%d日', time.localtime(time.time()))
    if mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';") is not None:
        duty_name=mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';")
        return duty_name[0]['duty_name']


    duty_date = time.strftime('%Y年%m月%#d日', time.localtime(time.time()))
    if mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';") is not None:
        duty_name=mysql_query_data(f"SELECT duty_name from duty_schedule WHERE date='{duty_date}';")
        return duty_name[0]['duty_name']

    duty_name = '鼎鼎'
    return duty_name

def excel_write_data(ws, value):
    global row_index
    ws.cell(row=row_index, column=col_index).value = value
    row_index = row_index + 1
    return ws


bad_increated_data = mysql_query_data(bad_increated_sql) if len(mysql_query_data(bad_increated_sql)) != 0 else [{'集群名称': '', '集群ID': '', '今日差距': ''}]
fault_sectors_data = mysql_query_data(fault_sectors_sql) if len(mysql_query_data(fault_sectors_sql)) != 0 else [{'集群名称': '', '集群ID': '', '惩罚算力': ''}]
orphan_block_data = mysql_query_data(orphan_block_sql) if len(mysql_query_data(orphan_block_sql)) != 0 else [{'集群名称': '', '集群ID': '', '孤块高度': '', '孤块原因': ''}]

logging.debug(bad_increated_data)
logging.debug(fault_sectors_data)
logging.debug(orphan_block_data)

bad_increated_title = [key for key in bad_increated_data[0].keys()]
bad_increated_title.append('原因')
fault_sectors_title = [key for key in fault_sectors_data[0].keys()]
fault_sectors_title.append('原因')
orphan_block_title = [key for key in orphan_block_data[0].keys()]

logging.debug(bad_increated_title)
logging.debug(fault_sectors_title)
logging.debug(orphan_block_title)


bad_increated_records = [[v for v in i.values()] for i in bad_increated_data]
fault_sectors_records = [[v for v in i.values()] for i in fault_sectors_data]
orphan_block_records = [[v for v in i.values()] for i in orphan_block_data]
logging.debug(bad_increated_records)
logging.debug(fault_sectors_records)
logging.debug(orphan_block_records)


if __name__ == '__main__':
    duty_name = get_duty_name()
    logging.debug(duty_name)
    work_book_name = f'运维交接表-{datetime.date.today()}_{duty_name}.xlsx'
    work_book_dir = './'
    work_book = work_book_dir + work_book_name
    wb = openpyxl.Workbook()
    ws = wb.active  # type:# openpyxl.worksheet.worksheet.Worksheet
    # ws = wb.active  -> openpyxl.worksheet.worksheet.Worksheet:
    ws.title = f'{work_sheet}'
    ws.index = 0

    ws.merge_cells(start_row=row_index, end_row=row_index, start_column=col_index, end_column=col_index + col_offset)
    ws.cell(row=row_index, column=col_index, value='').fill = PatternFill('solid', fgColor='7A7A7A')
    ws.cell(row=row_index, column=col_index, value='').font = Font(u'仿宋', size=18, bold=True, italic=False,
                                                                   strike=False, color=colors.BLACK)
    ws.cell(row=row_index, column=col_index, value='').alignment = Alignment(horizontal='center', vertical='center')
    excel_write_data(ws, '前班重要事宜')
    for i in range(1, 9):
        ws.merge_cells(start_row=row_index, end_row=row_index, start_column=col_index, end_column=col_index + col_offset)
        ws = excel_write_data(ws=ws, value=None)

    ws.merge_cells(start_row=row_index, end_row=row_index, start_column=col_index, end_column=col_index + col_offset)
    ws.cell(row=row_index, column=col_index, value='').fill = PatternFill('solid', fgColor='7A7A7A')
    ws.cell(row=row_index, column=col_index, value='').font = Font(u'仿宋', size=18, bold=True, italic=False,
                                                                   strike=False, color=colors.BLACK)
    ws.cell(row=row_index, column=col_index, value='').alignment = Alignment(horizontal='center', vertical='center')
    ws = excel_write_data(ws=ws, value='本班重要事宜')

    for i in range(1, 9):
        ws.merge_cells(start_row=row_index, end_row=row_index, start_column=col_index, end_column=col_index + col_offset)
        ws = excel_write_data(ws=ws, value=f'{i}、')

    ## 产量不达标
    ws.merge_cells(start_row=row_index, end_row=row_index, start_column=col_index, end_column=col_index + col_offset)
    ws.cell(row=row_index, column=col_index, value='').fill = PatternFill('solid', fgColor='7A7A7A')
    ws.cell(row=row_index, column=col_index, value='').font = Font(u'仿宋', size=18, bold=True, italic=False,
                                                                   strike=False, color=colors.BLACK)
    ws.cell(row=row_index, column=col_index, value='').alignment = Alignment(horizontal='center', vertical='center')
    ws = excel_write_data(ws=ws, value='产量不达标')
    ws.append(bad_increated_title)
    row_index =row_index +1
    bad_increated_record = [key for key in bad_increated_data[0].values()]
    for bad_increated_record in bad_increated_records:
        ws.append(bad_increated_record)
        row_index = row_index + 1


    ## 惩罚算力
    ws.merge_cells(start_row=row_index, end_row=row_index, start_column=col_index, end_column=col_index + col_offset)
    ws.cell(row=row_index, column=col_index, value='').fill = PatternFill('solid', fgColor='7A7A7A')
    ws.cell(row=row_index, column=col_index, value='').font = Font(u'仿宋', size=18, bold=True, italic=False,
                                                                   strike=False, color=colors.BLACK)
    ws.cell(row=row_index, column=col_index, value='').alignment = Alignment(horizontal='center', vertical='center')
    ws = excel_write_data(ws=ws, value='惩罚算力')
    ws.append(fault_sectors_title)
    row_index = row_index + 1
    for fault_sectors_record in fault_sectors_records:
        ws.append(fault_sectors_record)
        row_index = row_index + 1


    ## 孤块
    ws.merge_cells(start_row=row_index, end_row=row_index, start_column=col_index, end_column=col_index + col_offset)
    ws.cell(row=row_index, column=col_index, value='').fill = PatternFill('solid', fgColor='7A7A7A')
    ws.cell(row=row_index, column=col_index, value='').font = Font(u'仿宋', size=18, bold=True, italic=False,
                                                                   strike=False, color=colors.BLACK)
    ws.cell(row=row_index, column=col_index, value='').alignment = Alignment(horizontal='center', vertical='center')
    ws = excel_write_data(ws=ws, value='孤块')
    ws.append(orphan_block_title)
    row_index = row_index + 1
    for orphan_block_record in orphan_block_records:
        ws.append(orphan_block_record)
        row_index = row_index + 1


    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 150

    thin = Side(border_style='thin',color='000000')
    for row in range(1,row_index):
        for col in range(1,col_index+col_offset+1):
            ws.cell(row, col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.save(f'{work_book}')
    logging.info(f'生成{work_book}')

    # put excel to oss
    auth = oss2.Auth(access_key_id, access_key_secret)
    bucket = oss2.Bucket(auth, endpoint_name, bucket_name)

    work_book_oss_path = 'ops_duty' + '/' + time.strftime('%Y-%m', time.localtime(time.time())) + '/' +work_book_name
    logging.debug(work_book_oss_path)
    with open(work_book, 'rb') as fileobj:
        # fileobj.seek(1000, os.SEEK_SET)
        bucket.put_object(work_book_oss_path, fileobj)
    # https://ipfser-pro.oss-cn-hongkong.aliyuncs.com/ops_duty/%E8%BF%90%E7%BB%B4%E4%BA%A4%E6%8E%A5%E8%A1%A8-2023-08-23_%E5%A4%A9%E7%A9%BA%E6%98%9F.xlsx
    work_book_oss_url = 'https://' + bucket_name + '.' + endpoint_name + '/' + work_book_oss_path
    logging.info(work_book_oss_url)

    # send work_book oss_url to ding
    ding_webhook_url = 'https://oapi.dingtalk.com/robot/send?access_token=58af4c18b06991e6928b6d948ff7583e9383f730d777c24d64ba94d58b3dbcab'
    msg = {
        "msgtype": "text",
        "text": {
            "content": work_book_oss_url
        }
    }
    response = requests.post(ding_webhook_url, json=msg)
    if response.status_code != 200:
        logging.error('日报发送钉钉失败')
    logging.info('日报发送钉钉成功')